from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QApplication,QMainWindow, QMessageBox, QTableWidgetItem,QLCDNumber,QTableWidget,QDialog
from PyQt6.QtGui import QPalette,QDoubleValidator,QIcon

from rc_onesave_ui import Ui_MainWindow
from names import Ui_Dialog
from Chrome_conentrate_and_ore_cal import ChromeOreAnalysis
from FeroChrome_calculation import FeroChromeAnalysis
from Iron_calculation import IronAnalysis
import os
import sys
import time
import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph,Image,Spacer,PageBreak
from reportlab.lib.styles import ParagraphStyle, TA_LEFT

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch,cm
from reportlab.lib.styles import getSampleStyleSheet





class LabSystem(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.show()
        self.setWindowTitle("RCI Analytical Services")
        icon = QIcon("Pics/logo.ico")
        self.setWindowIcon(icon)

        self.KnownValue.currentChanged.connect(self.on_tab_changed)

        self.analysis=[] # put the three analysis here 
        self.init_analysis()
        self.current_sample_index = [0,0,0]
        self.currentSampleValues = [{},{},{}]

        self.calculated =[False,False,False]

        self.table_widgets=[]
        self.init_table_widgets()

        self.values = {}
        self.labels = {}
        self.init_values_labels()

        self.sample_values = {}
        self.sample_cal_labels = [self.cr_SampleCalculations,self.feSampletabletext,self.SampleCalculations]
        self.splitters =[self.splitter_7,self.splitter_2,self.splitter_5]
        self.init_sample_values_labels()
        self.init_edit()

        self.connect_edit_buttons()

        self.sample_info_labels = [self.cr_sample_info,self.FeSampleInfo,self.label]

        self.factor_next_buttons = [self.cr_factor_next_button,self.fe_factor_next_button,self.iron_factor_next_button]

        self.sample_table_widgets = [self.cr_sample_tableWidget,self.fe_tableWidget,self.IronSampleTable]
        for table in self.sample_table_widgets:
            table.cellChanged.connect(self.sample_item_changed)

        self.sample_next_buttons = [self.cr_sample_next_button,self.fe_sample_next_button,self.iron_sample_next_button]

        self.factor_displays = [self.cr_factor_display,self.fe_factor_display,self.factor_display]
        for display in self.factor_displays:
            display.setSegmentStyle(QLCDNumber.SegmentStyle.Flat)
            display.setStyleSheet("QLCDNumber { color: white;background-color:red }")



        self.index = 0
        self.init_tab()

        self.connect_next_buttons()
        self.save_button.clicked.connect(self.openNamesDialog)
        
        self.set_on_text_changed()


    def on_tab_changed(self,index):
        self.index = index
        if self.current_sample_index[self.index] ==0:
            self.init_tab()
        # if  self.calculated[self.index] == False: # need to add condition of not cleared
        #     self.init_tab() 


    def init_tab(self):
        self.current_sample_index[self.index] = 0
        self.currentSampleValues[self.index] = {}
        self.populate_known_samples_table()
        self.update_sample_info_label()
        self.hide_sample_calculations()
        self.hide_edit()
        self.populate_comboboxes()



    def init_analysis(self):
        self.analysis.append(ChromeOreAnalysis())
        self.analysis.append(FeroChromeAnalysis())
        self.analysis.append(IronAnalysis())

    def init_table_widgets(self):
        self.table_widgets.append(self.cr_factor_TableWidget)
        self.table_widgets.append(self.FeFactorTableWidget)
        self.table_widgets.append(self.IronTableWidget)


    def init_values_labels(self):
        self.values['grams']=[self.cr_factor_grams_value, self.fe_factor_grams_value,self.factor_grams_value]  
        self.values['ml'] = [self.cr_factor_ml_value,self.fe_factor_ml_value,self.factor_ml_value]     

        self.labels['grams'] =[self.cr_factor_grams_text,self.fe_factor_grams_text,self.iron_factor_grams_text]
        self.labels['ml']=[self.cr_factor_ml_text,self.fe_factor_ml_text,self.iron_factor_ml_text]


    def init_sample_values_labels(self):
        self.sample_values['ref_id']=[self.cr_sample_ref_id_value,self.fe_sample_ref_id,self.sample_ref_id]
        self.sample_values['grams']=[self.cr_sample_grams_value,self.fe_sample_grams_value,self.sample_grams]
        self.sample_values['ml']=[self.cr_sample_ml_value,self.fe_sample_ml_value,self.sample_ml]
    

    def init_edit(self):
        self.edit_labels = [self.cr_edit_label,self.fe_edit_label,self.iron_edit_label]
        self.sample_comboboxes = [self.cr_sample_combobox,self.fe_sample_combobox,self.iron_sample_combobox]
        self.input_comboboxes = [self.cr_input_combobox,self.fe_input_combobox,self.iron_input_combobox]
        self.edit_inputs = [self.cr_edit_input,self.fe_edit_input,self.iron_edit_input]
        self.edit_buttons = [self.ce_edit_button,self.fe_edit_button,self.iron_edit_button]


    def connect_next_buttons(self):
        self.cr_factor_next_button.clicked.connect(self.factor_results)
        self.cr_sample_next_button.clicked.connect(self.sample_results) 
        self.fe_factor_next_button.clicked.connect(self.factor_results)
        self.fe_sample_next_button.clicked.connect(self.sample_results)
        self.iron_factor_next_button.clicked.connect(self.factor_results)
        self.iron_sample_next_button.clicked.connect(self.sample_results)

    def connect_edit_buttons(self):
        for button in self.edit_buttons:
            button.clicked.connect(self.editSample)

        
    def set_on_text_changed(self):
        self.cr_factor_grams_value.textChanged.connect(self.update_grams_field)
        self.fe_factor_grams_value.textChanged.connect(self.update_grams_field)
        self.factor_grams_value.textChanged.connect(self.update_grams_field)

        self.cr_factor_ml_value.textChanged.connect(self.update_ml_field)
        self.fe_factor_ml_value.textChanged.connect(self.update_ml_field)
        self.factor_ml_value.textChanged.connect(self.update_ml_field)


    def sample_item_changed(self,row,col):
        if col in (0, 1, 2):  # Check if edited column is 1 (grams) or 2 (ml) # add 0 also to edit ref id in 
            self.update_row_values(row)

    def update_row_values(self,row):
        try:
            try:
                sample_id = self.sample_table_widgets[self.index].item(row, 0).text()
                grams_text = self.sample_table_widgets[self.index].item(row, 1).text()
                ml_text = self.sample_table_widgets[self.index].item(row, 2).text()
            except: # means table is not populated yet so no need to go through this function..
                return

            grams_float = float(grams_text)
            ml_float = float(ml_text)
            edit = True
            
            updated_samples = self.analysis[self.index].add_and_calculate_sample(sample_id, grams_float, ml_float,edit,row)
            last_entry = updated_samples[row]

            self.sample_table_widgets[self.index].setItem(row, 3, QTableWidgetItem(str(last_entry[3])))  # cal_percent_cr
            # Make columns 3 and 4 non-editable
            # self.sample_table_widgets[self.index].item(row, 3).setFlags(Qt.ItemFlag.ItemIsEnabled)

            if self.index != 1:
                self.sample_table_widgets[self.index].setItem(row, 4, QTableWidgetItem(str(last_entry[4])))  


        except ValueError:
            QMessageBox.warning(self, "Input Error", "Please enter valid numbers for grams and ml.")


    def editSample(self):  
        try:
            selected_sample = self.sample_comboboxes[self.index].currentText() 
            selected_input = self.input_comboboxes[self.index].currentText() 
            value = float(self.edit_inputs[self.index].text())

        except:

            QMessageBox.warning(self, "Input Error", "Please fill in all Edit Fields with valid values.")
            return
        
        row = self.sample_comboboxes[self.index].currentIndex()  # Get the index of the selected item
        if selected_input =="grams":
            grams = value
            self.table_widgets[self.index].setItem(row, 1, QTableWidgetItem(str(value)))
            ml = float(self.table_widgets[self.index].item(row, 2).text())

        else:
            ml = value
            self.table_widgets[self.index].setItem(row, 2, QTableWidgetItem(str(value)))
            grams = float(self.table_widgets[self.index].item(row, 1).text())

        self.currentSampleValues[self.index][selected_sample] = (grams, ml)
        anaylsis_results = self.analysis[self.index].calculate_factors(self.currentSampleValues[self.index])
        self.display_results_in_table(anaylsis_results)
            
        bias_violation = False
        for result in anaylsis_results:
            if abs(result[6]) > 0.5:
                QMessageBox.warning(self, "Bias Violation", f"Sample '{result[0]}' exceeded the bias tolerance with a bias of {result[6]}.")
                bias_violation = True
                break
        if not bias_violation:
            self.calculated[self.index]= True
            self.show_sample_calculations() 
            self.update_factor_display()
            self.hide_edit()
        
        self.edit_inputs[self.index].clear()

        
    def showError(self, message):
        QMessageBox.critical(self, "Error", message)
        
        
    def populate_known_samples_table(self):
        self.table_widgets[self.index].setRowCount(len(self.analysis[self.index].known_samples))
        for row, sample_name in enumerate(self.analysis[self.index].known_samples.keys()):
            self.table_widgets[self.index].setItem(row, 0, QTableWidgetItem(sample_name))
            self.table_widgets[self.index].item(row, 0).setFlags(Qt.ItemFlag.ItemIsEnabled)
            value = self.analysis[self.index].known_values[row]
            self.table_widgets[self.index].setItem(row, 5, QTableWidgetItem(str(value)))
            self.table_widgets[self.index].item(row, 5).setFlags(Qt.ItemFlag.ItemIsEnabled)

    
    
    def change_next_into_clear_button(self):
        self.factor_next_buttons[self.index].setText("Clear")
        try:
            self.factor_next_buttons[self.index].clicked.disconnect()
        except TypeError:
            # No connections exist yet, so ignore
            pass
        # Connect the clear functionality
        self.factor_next_buttons[self.index].clicked.connect(self.clear_all_data)

                
    def update_sample_info_label(self):
        if self.current_sample_index[self.index] < len(self.analysis[self.index].known_samples):
            sample_name = list(self.analysis[self.index].known_samples.keys())[self.current_sample_index[self.index]]
            self.sample_info_labels[self.index].setText(f"Current Sample: {sample_name}")
        else:
            self.sample_info_labels[self.index].setText("All samples processed.")
            
    def factor_results(self): 
        # self.check_not_empty() # check texts are not empty first
        try:
            grams = float(self.values['grams'][self.index].text())
            ml = float(self.values['ml'][self.index].text())
        except:

            # Validate input before proceeding
            # if not grams or not ml or not known_value:
            QMessageBox.warning(self, "Input Error", "Please fill in all fields with valid values.")
            return

        sample_name = list(self.analysis[self.index].known_samples.keys())[self.current_sample_index[self.index]]
        self.currentSampleValues[self.index][sample_name] = (grams, ml)#, known_value)

        # Prepare for the next sample or finish
        self.current_sample_index[self.index] += 1
        if self.current_sample_index[self.index] >= len(self.analysis[self.index].known_samples):
            # Since calculate_factors expects a dictionary, now we pass self.currentSampleValues directly
            chrom_calculated_values = self.analysis[self.index].calculate_factors(self.currentSampleValues[self.index])
            self.display_results_in_table(chrom_calculated_values)
            
            bias_violation = False
            for result in chrom_calculated_values:
            # Assuming the bias is the last element in the result list
                if abs(result[6]) > 0.5:
                    QMessageBox.warning(self, "Bias Violation", f"Sample '{result[0]}' exceeded the bias tolerance with a bias of {result[6]}.")
                    bias_violation = True
                    break
            if not bias_violation:
                self.calculated[self.index]= True
                self.show_sample_calculations() # show the second step table
                self.update_factor_display()
            else:
                self.show_edit()
                # self.populate_comboboxes()


            
            self.hide_inputs_and_calculate()
            self.change_next_into_clear_button()

        else:
            self.update_sample_info_label()

        self.values['grams'][self.index].clear()
        self.values['ml'][self.index].clear()
    
    def display_results_in_table(self, results):
        self.table_widgets[self.index].setRowCount(len(results))  # Assuming resultsTable is your QTableWidget
        for row, sample in enumerate(results):
            for col, data in enumerate(sample):
                self.table_widgets[self.index].setItem(row, col, QTableWidgetItem(str(data)))
                self.table_widgets[self.index].item(row, col).setFlags(Qt.ItemFlag.ItemIsEnabled)


    def update_factor_display(self):
        factor_average = self.analysis[self.index].factor_average
        self.factor_displays[self.index].display(factor_average)

    def clear_display(self):
        self.factor_displays[self.index].display(0)

            
    def hide_inputs_and_calculate(self):
        # Hide UI elements
        self.labels['grams'][self.index].hide()
        self.values['grams'][self.index].hide()
        self.labels['ml'][self.index].hide()
        self.values['ml'][self.index].hide()
    
    def reset_current_state(self):
        self.init_analysis()
        self.currentSampleValues[self.index] = {}
        self.current_sample_index[self.index] = 0




    
    def show_all_factor_text_value(self):
        self.labels['grams'][self.index].show()
        self.values['grams'][self.index].show()
        self.labels['ml'][self.index].show()
        self.values['ml'][self.index].show()
        
    def clear_all_data(self):
   
        self.reset_current_state()
        self.show_all_factor_text_value()
  
        self.factor_next_buttons[self.index].setText("Next")
        self.factor_next_buttons[self.index].clicked.disconnect()
        self.factor_next_buttons[self.index].clicked.connect(self.factor_results)
        # Optionally, clear any displayed results in your table
        self.table_widgets[self.index].clearContents() 
        self.table_widgets[self.index].setRowCount(0)
        # Reset the UI to its initial state if needed
        
        self.sample_table_widgets[self.index].clearContents()
        self.sample_table_widgets[self.index].setRowCount(0)
        
        self.populate_known_samples_table()
        self.update_sample_info_label()
        self.hide_sample_calculations()

        self.clear_display()
        self.calculated[self.index ] = False
        self.hide_edit()
        # self.populate_comboboxes()
    
    def sample_results(self):
        # Collect the input data
        sample_id = self.sample_values['ref_id'][self.index].text()  # Assumes sampleIdLineEdit is your QLineEdit for sample ID
        grams = self.sample_values['grams'][self.index].text()  # Assumes gramsLineEdit is for grams input
        ml = self.sample_values['ml'][self.index].text()  # Assumes mlLineEdit is for ml input

        # Convert inputs to appropriate types (assuming grams and ml should be floats)
        try:
            grams_float = float(grams)
            ml_float = float(ml)
        except ValueError:
            # Handle invalid input (e.g., show an error message)
            QMessageBox.warning(self, "Input Error", "Please enter valid numbers for grams and ml.")
            return

        # Call add_and_calculate_sample with the collected data
        try:
            updated_samples = self.analysis[self.index].add_and_calculate_sample(sample_id, grams_float, ml_float)
        except ValueError as e:
            # Handle the case where factor_average has not been calculated
            QMessageBox.warning(self, "Calculation Error", str(e))
            return

        # Assuming the last entry in updated_samples is the one we just added
        last_entry = updated_samples[-1]
        # The expected format of last_entry is [ref_id, grams, ml, cal_percent_cr, calc_cr2o3]

        # Add the new entry to the table
        rowPosition = self.sample_table_widgets[self.index].rowCount()
        self.sample_table_widgets[self.index].insertRow(rowPosition)
        for col, item in enumerate(last_entry):
            self.sample_table_widgets[self.index].setItem(rowPosition, col, QTableWidgetItem(str(item)))
        # make them not editable
        self.sample_table_widgets[self.index].item(rowPosition, 3).setFlags(Qt.ItemFlag.ItemIsEnabled)
        if self.index != 1:
            self.sample_table_widgets[self.index].item(rowPosition, 4).setFlags(Qt.ItemFlag.ItemIsEnabled)  # Column 4 (calc_cr2o3)




        # Optionally, clear the input fields after adding the data to the table
        self.sample_values['ref_id'][self.index].clear()
        self.sample_values['grams'][self.index].clear()
        self.sample_values['ml'][self.index].clear()
        


    def hide_edit(self):
        self.edit_labels[self.index].setVisible(False)
        self.sample_comboboxes[self.index].setVisible(False)
        self.input_comboboxes[self.index].setVisible(False)
        self.edit_inputs[self.index].setVisible(False)
        self.edit_buttons[self.index].setVisible(False)

    

    def show_edit(self):
        self.edit_labels[self.index].setVisible(True)
        self.sample_comboboxes[self.index].setVisible(True)
        self.input_comboboxes[self.index].setVisible(True)
        self.edit_inputs[self.index].setVisible(True)
        self.edit_buttons[self.index].setVisible(True)

    def populate_comboboxes(self):

        sample_names = list(self.analysis[self.index].known_samples.keys())
        self.sample_comboboxes[self.index].clear()
        self.input_comboboxes[self.index].clear()
        self.sample_comboboxes[self.index].addItems(sample_names)
        self.input_comboboxes[self.index].addItems(["grams","mL"])


    def hide_sample_calculations(self):
        self.sample_cal_labels[self.index].setVisible(False)
        self.sample_table_widgets[self.index].setVisible(False)
        self.splitters[self.index].setVisible(False)

        if self.calculated[0] or self.calculated[1] or self.calculated[2]:
            self.save_button.setVisible(True)

        else:
            self.save_button.setVisible(False)


    def show_sample_calculations(self):
        self.sample_cal_labels[self.index].setVisible(True)
        self.sample_table_widgets[self.index].setVisible(True)
        # self.sample_save_buttons[self.index].setVisible(True)
        self.splitters[self.index].setVisible(True)

        # if self.calculated[0] and not(self.calculated[1]) and not(self.calculated[2]):
        if self.calculated[0] or self.calculated[1] or self.calculated[2]:
            self.save_button.setVisible(True)

    def update_grams_field(self):
        text = self.values['grams'][self.index].text()
        self.table_widgets[self.index].setItem(self.current_sample_index[self.index], 1, QTableWidgetItem(text))
        try:
            self.table_widgets[self.index].item(self.current_sample_index[self.index], 1).setFlags(Qt.ItemFlag.ItemIsEnabled)
        except:
            pass


        

    def update_ml_field(self):
        text = self.values['ml'][self.index].text()
        self.table_widgets[self.index].setItem(self.current_sample_index[self.index], 2, QTableWidgetItem(text))
        try:
            self.table_widgets[self.index].item(self.current_sample_index[self.index], 2).setFlags(Qt.ItemFlag.ItemIsEnabled)
        except:
            pass


    def extractSampleTables(self):
        sample_data=[[],[],[]]
        for i in range(3):
            for row in range(self.sample_table_widgets[i].rowCount()):
                row_data = []
                for col in range(self.sample_table_widgets[i].columnCount()):
                    item = self.sample_table_widgets[i].item(row, col)
                    row_data.append(item.text() if item else "")
                sample_data[i].append(row_data)
            # print(sample_data[i])
        return sample_data

    def findSampleIndex(self,ref_id, table):
        for i in range(len(table)):
            if table[i][0] == ref_id:
                return i
            
        return -1

    def openNamesDialog(self):
        from namesdialog import NamesDialog
        dialog = NamesDialog(self)
        result = dialog.exec()
        print(result)

        if result:
            analyst_name = dialog.ui.analyst_name.text()
            supervisor_name = dialog.ui.supervisor_name.text()
            self.savePdfandSheet(analyst_name,supervisor_name)

    def savePdfandSheet(self,analyst_name,supervisor_name):
        
        file_time = time.strftime("Date_%d-%m-%Y_Time_%H-%M-%S")
        image_path = os.path.join("Pics/rci as logo.png")
        samples_data = [["Sample Ref ID","Cr %","Cr2O3 %","Fe %", "FeO %"]]
        # samples_data = [["Sample ","CR %","Cr2O3 %","Fe %", "Fe0 %"]]
        sampletables = self.extractSampleTables()
        cr_fe_data =[["Sample Ref ID","Cr %","Fe %", "Cr/Fe Ratio"]]
        #iron chrome
        i=0
        # for i in range(len(sampletables[2])):
        while i < len(sampletables[2]):
            index_chrome = self.findSampleIndex(sampletables[2][i][0],sampletables[0])
            if index_chrome != -1 :
                item = [sampletables[2][i][0],"",sampletables[0][index_chrome][4],"",sampletables[2][i][4]]
                # ratio_item = [sampletables[2][i][0],sampletables[0][index_chrome][3],sampletables[2][i][3],str(round(float(sampletables[0][index_chrome][3])/float(sampletables[2][i][3]),2))]
                # cr_fe_data.append(ratio_item)
                
                # item = ["Iron + Chrome","",sampletables[0][index_chrome][4],"",sampletables[2][i][4]]

                samples_data.append(item)
                sampletables[2].pop(i)
                sampletables[0].pop(index_chrome)
                i-=1
            i+=1

        #iron fe
        i=0
        # for i in range(len(sampletables[2])):
        while i < len(sampletables[2]):
            index_fe = self.findSampleIndex(sampletables[2][i][0],sampletables[1])
            if index_fe != -1 :
                item = [sampletables[2][i][0],sampletables[1][index_fe][3],"",sampletables[2][i][3],""]
                # item = ["Iron + Ferro chrome",sampletables[1][index_fe][3],"",sampletables[2][i][3],""]
                ratio_item = [sampletables[2][i][0],sampletables[1][index_fe][3],sampletables[2][i][3],str(round(float(sampletables[1][index_fe][3])/float(sampletables[2][i][3]),2))]
                cr_fe_data.append(ratio_item)


                samples_data.append(item)
                sampletables[2].pop(i)
                sampletables[1].pop(index_fe)
                i-=1
            i+=1

        # i=0
        for i in range(len(sampletables[0])):
            item = [sampletables[0][i][0],sampletables[0][i][3],sampletables[0][i][4],"",""]
            # item = ["Chrome",sampletables[0][i][3],sampletables[0][i][4],"",""]
            samples_data.append(item)

        for i in range(len(sampletables[1])):
            item = [sampletables[1][i][0],sampletables[1][i][3],"","",""]
            # item = ["Ferro Chrome",sampletables[1][i][3],"","",""]

            samples_data.append(item)

        for i in range(len(sampletables[2])):
            item = [sampletables[2][i][0],"","",sampletables[2][i][3],sampletables[2][i][4]]
            # item = ["Iron","","",sampletables[2][i][3],sampletables[2][i][4]]

            samples_data.append(item)

                

        print(samples_data)

        current_date = datetime.date.today()
        current_time = datetime.datetime.now().time()
        # try:

        self.saveAllTablesPdf(file_time,image_path,samples_data,cr_fe_data,current_date,current_time,analyst_name,supervisor_name)
        self.saveExcel(file_time,image_path,samples_data,current_date,current_time)

        QMessageBox.information(self, "Success", f"Data saved successfully to  rci_{file_time}.pdf & rci_{file_time}.xlsx !")

        # except:
        #     QMessageBox.warning(self, "Error in Saving PDF", "An Error Occured during Saving the file , please try again later")


    def saveAllTablesPdf(self,file_time,image_path,samples_data,ratio_data,current_date,current_time,analyst_name,supervisor_name):

        filename = "rci"
        ext=".pdf"
        final_name = f"{filename}_{file_time}{ext}"

        doc = SimpleDocTemplate(final_name, pagesize=letter, leftMargin=20,  # Adjust as needed
                        rightMargin=2.2, topMargin=1.5, bottomMargin=2.5)

        elements = []

        # Create content elements
        image_width = 2.5  # Adjust width in inches
        image_height = 0.5  # Adjust height in inches
        elements.append(Image(image_path, width=inch * image_width, height=inch * image_height))


        header_font_size = 10

        styles = getSampleStyleSheet()
        header_style = styles['Heading1']
        header_style.fontSize = header_font_size
        header_style.alignment =1

        elements.append(Spacer(1, 0.5 * cm))  # Adjust spacing

        header_text = "FINAL CERTIFICATE OF ANALYSIS"
        elements.append(Paragraph(header_text, header_style))
        header_text = "REVISION 0"
        elements.append(Paragraph(header_text, header_style))


        # Define styles (optional)
        styles = getSampleStyleSheet()
        contact_style = styles['Normal']  # Adjust style as needed

        # Create paragraphs for contact information
        from_text = Paragraph("FROM:         RCI Analytical Services", contact_style)
        # tel_text = Paragraph("TEL:", contact_style)
        # fax_text = Paragraph("FAX:", contact_style)
        date_text = Paragraph(f"Date:       {current_date}", contact_style)
        time_text = Paragraph(f"Time:       {current_time.hour:02d}:{current_time.minute:02d}:{current_time.second:02d}", contact_style)

        # elements.append(Spacer(1, 0.5 * cm))  # Adjust spacing

        elements.append(from_text)
        # elements.append(tel_text)
        # elements.append(fax_text)
        elements.append(date_text)
        elements.append(time_text)

        elements.append(Spacer(1, 0.5 * cm))  # Adjust spacing

        analyst_text    = Paragraph(f"Analyst:        {analyst_name}",contact_style)
        supervisor_text = Paragraph(f"Supervisor:     {supervisor_name}",contact_style)

        elements.append(analyst_text)
        elements.append(supervisor_text)

        elements.append(Spacer(1, 0.5 * cm))  # Adjust spacing

        sample_table = Table(samples_data, hAlign='LEFT')
        sample_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.blue),
            ('BACKGROUND', (0,0), (0,-1), colors.blue),

            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('TEXTCOLOR',(0,0),(0,-1),colors.whitesmoke)

        ]))
        elements.append(Paragraph("<b>Samples Table</b>", style=ParagraphStyle(
                alignment=TA_LEFT,  
                fontSize=24,   
                fontName="Helvetica-Bold",
                name='FactorCalculationTable'
            )))
        elements.append(Table([[""]], colWidths=[doc.width]))

        elements.append(sample_table)

        elements.append(Table([[""]], colWidths=[doc.width]))
        ratio_table = Table(ratio_data, hAlign='LEFT')
        ratio_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.blue),
            ('BACKGROUND', (0,0), (0,-1), colors.blue),

            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('TEXTCOLOR',(0,0),(0,-1),colors.whitesmoke)

        ]))

        elements.append(Paragraph("<b>Cr/Fe Ratio Table</b>", style=ParagraphStyle(
                        alignment=TA_LEFT,  
                        fontSize=24,   
                        fontName="Helvetica-Bold",
                        name='FactorCalculationTable'
                    )))
        elements.append(Table([[""]], colWidths=[doc.width]))

        elements.append(ratio_table)

        
        elements.append(PageBreak())





        factors_data = [[["Sample Name", "Grams", "mL", "Factor", "%Cr", "Known %", "Bias",]],
            [["Sample Name", "Grams", "mL", "Factor", "%Cr", "Known %", "Bias"]],
            [["Sample Name", "Grams", "mL", "Factor", "%Fe", "Known %", "Bias", "%FeO"]]]

        for i in range(3):
            for row in range(self.table_widgets[i].rowCount()):
                # names_of_samples = list(self.analysis[self.index].known_samples.keys())
                # row_data = [names_of_samples[row]]  # Start each row with the sample name
                row_data=[]
                for col in range(self.table_widgets[i].columnCount()):
                    item = self.table_widgets[i].item(row, col)
                    row_data.append(item.text() if item else "")
                factors_data[i].append(row_data)

            if (self.calculated[i]):
                # Add the Factors Table to the elements
                factors_table = Table(factors_data[i], hAlign='LEFT')
                factors_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.blue),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                    ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)
                ]))

                
                elements.append(Paragraph(f"<b>{self.analysis[i].name} Factor Calculation Table</b>", style=ParagraphStyle(
                                alignment=TA_LEFT,  
                                fontSize=24,   
                                fontName="Helvetica-Bold",
                                name='FactorCalculationTable'
                            )))
                elements.append(Table([[""]], colWidths=[doc.width]))

                elements.append(factors_table)
                
                # Adding a space between tables
                elements.append(Table([[""]], colWidths=[doc.width]))

                factor_average = self.analysis[i].factor_average
                elements.append(Paragraph(f"<b>Factor Average: {factor_average:.7f}</b>", style=ParagraphStyle(
                    alignment=TA_LEFT,  
                    fontSize=10,   
                    fontName="Helvetica-Bold",
                    name='FactorAverage'
                )))


                standard_deviation = self.analysis[i].standard_deviation
                elements.append(Paragraph(f"<b>Factor Standard Deviation: {standard_deviation:.7f}</b>", style=ParagraphStyle(
                    alignment=TA_LEFT,  
                    fontSize=10,   
                    fontName="Helvetica-Bold",
                    name='StandardDeviation'
                )))

                coefficient_of_variation = self.analysis[i].coefficient_of_variation
                elements.append(Paragraph(f"<b>Factor Standard Deviation Percentage: {coefficient_of_variation:.7f}</b>", style=ParagraphStyle(
                    alignment=TA_LEFT,  
                    fontSize=10,   
                    fontName="Helvetica-Bold",
                    name= 'StandardDeviationPercentage'
                )))

                elements.append(Table([[""]], colWidths=[doc.width]))


        elements.append(PageBreak())

        sample_data = [[["Ref ID", "Grams", "mL", "Cal % Cr","% Calc Cr2O3","Factor"]],
                    [["Ref ID", "Grams", "mL", "Cal % Cr","Factor"]],
                    [["Ref ID", "Grams", "mL", "%Fe", "%FeO","Factor"]]]

        for i in range(3):
            for row in range(self.sample_table_widgets[i].rowCount()):
                row_data = []
                for col in range(self.sample_table_widgets[i].columnCount()):
                    item = self.sample_table_widgets[i].item(row, col)
                    row_data.append(item.text() if item else "")
                factor_average = f"{self.analysis[i].factor_average:.7f}"
                row_data.append(factor_average)
                sample_data[i].append(row_data)

            # Add the Sample Table to the elements
            sample_table = Table(sample_data[i], hAlign='LEFT')
            sample_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.blue),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)
            ]))

            elements.append(Paragraph(f"<b>{self.analysis[i].name} Sample Calculation Table</b>", style=ParagraphStyle(
                alignment=TA_LEFT,  
                fontSize=24,   
                fontName="Helvetica-Bold",
                name='FactorCalculationTable'
            )))
            elements.append(Table([[""]], colWidths=[doc.width]))            

            elements.append(sample_table)
            elements.append(Table([[""]], colWidths=[doc.width]))




        doc.build(elements)


    def saveExcel(self,file_time,image_path,samples_data,current_date,current_time):
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Font, PatternFill


        file_name = f"rci_{file_time}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        img = Image(image_path)
        ws.add_image(img, 'A1')


        ws['D10'] = 'FINAL CERTIFICATE OF ANALYSIS'
        ws['D11'] = 'REVISION 0'

        bold_font = Font(bold=True)
        ws['D10'].font = bold_font
        ws['D11'].font = bold_font


        ws['A15'] = 'FROM'
        # ws['A16'] = 'TEL'
        # ws['A17'] = 'FAX'
        ws['A18'] = 'Date'
        ws['A19'] = 'Time'


        ws['B15'] = 'RCI  Analytical Services'
        # ws['B16'] = 'TEL'
        # ws['B17'] = 'FAX'
        ws['B18'] = f"{current_date}"
        ws['B19'] = f'{current_time}'

        headers = samples_data[0]
        for col_index, header in enumerate(headers, start=1):
            # ws.cell(row=20, column=col_index).value = header
            cell = ws.cell(row=20, column=col_index)
            cell.value = header
            cell.fill = PatternFill(start_color="0080FE", end_color="0080FE", fill_type="solid")  # Blue background
            cell.font = Font(color="FFFFFF")  # White font color


        for row_index, row in enumerate(samples_data[1:]):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index+21, column=col_index)
                cell.value = value
                if col_index == 1:

                    cell.fill = PatternFill(start_color="0080FE", end_color="0080FE", fill_type="solid")  # Blue background
                    cell.font = Font(color="FFFFFF")  # White font color

        wb.save(file_name)

        print(f"Excel sheet saved successfully: {file_name}")

                        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = LabSystem()
    window.show()
    sys.exit(app.exec())