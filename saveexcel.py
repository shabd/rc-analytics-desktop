import openpyxl

def findSampleIndex(ref_id, table):
    for i in range(len(table)):
        if table[i][0] == ref_id:
            return i
        
    return -1

sample_data = [[["Ref ID", "Grams", "Ml", "Cal % CR","% Calc Cr2O3"]],
            [["Ref ID", "Grams", "Ml", "Cal % CR"]],
            [["Ref ID", "Grams", "Ml", "%Fe", "%FeO"]]]

chrome_sample=[
               ['Rcr1', '2.0', '30.0', '4.55', '6.64'], 
               ['Ricr1', '5.0', '60.0', '3.64', '5.31'], 
               ['Rcr2', '2.0', '40.0', '6.06', '8.86'], 
               ['Ricr2', '4', '60.0', '4.55', '6.64'], 
               ['Rcr3', '8.0', '8.0', '0.3', '0.44']]

fe_sample=[
           ['Rfe1', '8.0', '90.0', '3.32'], 
           ['Rife1', '3.0', '50.0', '4.92'], 
           ['Rfe2', '5.0', '70.0', '4.14'], 
           ['Rife2', '15.0', '55.0', '1.08'], 
           ['Rfe3', '12.0', '34.0', '0.84']]

iron_sample=[ 
             ['Ricr1', '2.0', '10.0', '2.72', '3.5'], 
             ['Ri1', '4.0', '3.0', '0.41', '0.52'], 
             ['Ri2', '5.0', '8.0', '0.87', '1.12'], 
             ['Rife1', '5.0', '40.0', '4.35', '5.6'], 
             ['Ri3', '3.0', '8.0', '1.45', '1.87'], 
             ['Ricr2', '4.0', '70.0', '9.52', '12.25'], 
             ['Ri4', '5.0', '59.0', '6.42', '8.26'], 
             ['Rife2', '34.0', '45.0', '0.72', '0.93'], 
             ['Ri5', '6.0', '6.0', '0.54', '0.7']]


sampletables = [chrome_sample,fe_sample,iron_sample]


for i in range(3):
    for row in range(len(sampletables[i])):
        row_data = []
        for col in range(len(sampletables[i][0])):
            item = sampletables[i][row][col]
            row_data.append(item if item else "")
        sample_data[i].append(row_data)
    print(sample_data[i])







samples_data = [["Sample Ref ID","CR %","Cr2O3","Fe %", "Fe0 %"]]
for i in range(len(iron_sample)):
    if i < len(iron_sample):
        index_chrome = findSampleIndex(iron_sample[i][0],chrome_sample)
        if index_chrome != -1 :
            item = [iron_sample[i][0],chrome_sample[index_chrome][3],chrome_sample[index_chrome][4],iron_sample[i][3],iron_sample[i][3]]
            samples_data.append(item)
            iron_sample.pop(i)
            chrome_sample.pop(index_chrome)
            i+=1
        
for i in range(len(iron_sample)):
    if i < len(iron_sample):
        index_fe = findSampleIndex(iron_sample[i][0],fe_sample)
        if index_fe != -1 :
            item = [iron_sample[i][0],fe_sample[index_fe][3],"",iron_sample[i][3],iron_sample[i][3]]
            samples_data.append(item)
            iron_sample.pop(i)
            fe_sample.pop(index_fe)
            i+=1

for i in range(len(chrome_sample)):
    item = [chrome_sample[i][0],chrome_sample[i][3],chrome_sample[i][4],"",""]
    samples_data.append(item)

for i in range(len(fe_sample)):
    item = [fe_sample[i][0],fe_sample[i][3],"","",""]
    samples_data.append(item)

for i in range(len(iron_sample)):
    item = [iron_sample[i][0],"","",iron_sample[i][3],iron_sample[i][4]]
    samples_data.append(item)







import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active

# Load your image (replace 'logo.png' with your image file)
img = Image("Pics/rci as logo.png")

# Add the image to a specific cell (e.g., A1)
ws.add_image(img, 'A1')


ws['D10'] = 'FINAL CERTIFICATE OF ANALYSIS'
ws['D11'] = 'REVISION 0'

bold_font = Font(bold=True)
ws['D10'].font = bold_font
ws['D11'].font = bold_font


ws['A15'] = 'FROM'
ws['A16'] = 'TEL'
ws['A17'] = 'FAX'
ws['A18'] = 'Date'
ws['A19'] = 'Time'


ws['B15'] = 'RCI  Analytical Services'
# ws['B16'] = 'TEL'
# ws['B17'] = 'FAX'
ws['B18'] = 'passed Date'
ws['B19'] = 'passed Time'

# # Merge cells for main header (adjust row and column numbers)
# ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))  # Assuming headers span all columns

# # Set main header text and formatting (adjust font size and alignment)
# ws.cell(row=2, column=1).value = "FINAL CERTIFICATE OF ANALYSIS"
# # ws.cell(row=2, column=1).font.size = 10
# ws.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

# # Optional: Add a sub-header (adjust row and column)
# ws.cell(row=3, column=1).value = "REVISION 0"
# ws.cell(row=3, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')


# # Write headers (assuming first row of samples_data contains headers)
headers = samples_data[0]
for col_index, header in enumerate(headers, start=1):
    # ws.cell(row=20, column=col_index).value = header
    cell = ws.cell(row=20, column=col_index)
    cell.value = header
    cell.fill = PatternFill(start_color="0080FE", end_color="0080FE", fill_type="solid")  # Blue background
    cell.font = Font(color="FFFFFF")  # White font color


# Write data rows (starting from row 2)
for row_index, row in enumerate(samples_data[1:]):
    for col_index, value in enumerate(row, start=1):
        cell = ws.cell(row=row_index+21, column=col_index)
        cell.value = value
        if col_index == 1:
            cell.fill = PatternFill(start_color="0080FE", end_color="0080FE", fill_type="solid")  # Blue background
            cell.font = Font(color="FFFFFF")  # White font color

        # ws.cell(row=row_index+21, column=col_index).value = value


# Save the Excel sheet
filename = "extracted_data.xlsx"
wb.save(filename)

print(f"Excel sheet saved successfully: {filename}")
